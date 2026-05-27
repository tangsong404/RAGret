from __future__ import annotations

import logging
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

logger = logging.getLogger(__name__)

_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_OPENPYXL_READ_ERRORS = (TypeError, ValueError, KeyError, OSError, zipfile.BadZipFile)


class UnreadableXlsxError(Exception):
    """Raised when an .xlsx cannot be read by openpyxl or the XML fallback."""


def preprocess_xlsx(path: Path) -> str:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl is required for .xlsx preprocessing") from e

    try:
        return _preprocess_xlsx_openpyxl(path, openpyxl)
    except _OPENPYXL_READ_ERRORS as e:
        logger.warning("openpyxl failed for %s (%s); trying XML fallback", path.name, e)
        try:
            text = _preprocess_xlsx_xml_fallback(path)
        except Exception as fb_err:
            raise UnreadableXlsxError(str(path)) from fb_err
        if not text.strip():
            raise UnreadableXlsxError(str(path)) from e
        return text


def _preprocess_xlsx_openpyxl(path: Path, openpyxl: object) -> str:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)  # type: ignore[attr-defined]
    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                if any(c.strip() for c in cells):
                    rows.append("\t".join(cells))
            if rows:
                parts.append(f"--- sheet {sheet_name} ---\n" + "\n".join(rows))
    finally:
        wb.close()
    return "\n\n".join(parts).strip()


def _preprocess_xlsx_xml_fallback(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        shared = _read_shared_strings(zf)
        parts: list[str] = []
        for sheet_name, sheet_path in _sheet_xml_paths(zf):
            rows = _read_sheet_rows(zf, sheet_path, shared)
            if rows:
                parts.append(f"--- sheet {sheet_name} ---\n" + "\n".join(rows))
    return "\n\n".join(parts).strip()


def _read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    name = "xl/sharedStrings.xml"
    if name not in zf.namelist():
        return []
    root = ET.fromstring(zf.read(name))
    out: list[str] = []
    for si in root.findall(f"{{{_MAIN_NS}}}si"):
        chunks = [t.text or "" for t in si.iter(f"{{{_MAIN_NS}}}t")]
        out.append("".join(chunks))
    return out


def _sheet_xml_paths(zf: zipfile.ZipFile) -> list[tuple[str, str]]:
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_target: dict[str, str] = {}
    for rel in rel_root.findall(f"{{{_REL_NS}}}Relationship"):
        rid = rel.get("Id")
        target = rel.get("Target")
        if rid and target:
            rid_to_target[rid] = target.replace("\\", "/")

    out: list[tuple[str, str]] = []
    for sheet in wb_root.findall(f"{{{_MAIN_NS}}}sheets/{{{_MAIN_NS}}}sheet"):
        name = sheet.get("name") or "sheet"
        rid = sheet.get(f"{{{_OFFICE_REL_NS}}}id")
        target = rid_to_target.get(rid or "", "")
        if not target:
            continue
        if target.startswith("/"):
            sheet_path = target.lstrip("/")
        elif target.startswith("xl/"):
            sheet_path = target
        else:
            sheet_path = f"xl/{target}"
        out.append((name, sheet_path))
    return out


def _read_sheet_rows(zf: zipfile.ZipFile, sheet_path: str, shared: list[str]) -> list[str]:
    if sheet_path not in zf.namelist():
        return []
    root = ET.fromstring(zf.read(sheet_path))
    rows: list[str] = []
    for row in root.findall(f".//{{{_MAIN_NS}}}sheetData/{{{_MAIN_NS}}}row"):
        cells: list[str] = []
        for cell in row.findall(f"{{{_MAIN_NS}}}c"):
            cells.append(_cell_text(cell, shared))
        if any(c.strip() for c in cells):
            rows.append("\t".join(cells))
    return rows


def _cell_text(cell: ET.Element, shared: list[str]) -> str:
    cell_type = cell.get("t")
    value_el = cell.find(f"{{{_MAIN_NS}}}v")
    if value_el is None or value_el.text is None:
        inline = cell.find(f"{{{_MAIN_NS}}}is")
        if inline is not None:
            return "".join(t.text or "" for t in inline.iter(f"{{{_MAIN_NS}}}t"))
        return ""
    raw = value_el.text
    if cell_type == "s":
        try:
            return shared[int(raw)]
        except (ValueError, IndexError):
            return raw
    return raw
